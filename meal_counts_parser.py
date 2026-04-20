"""
Digital Meal Counts Parser — Imports Board meal count data from Excel
and stores in the database. Tracks daily counts by meal service (B/L/D),
prior year actuals, forecasts, admission groups, and special groups.
"""

import sqlite3
from datetime import date, datetime
from pathlib import Path

import openpyxl


# ─── Row map for each weekly sheet ───
_ROW_MAP = {
    "dates": 3,
    "py_breakfast": 5,
    "py_lunch": 6,
    "py_dinner": 7,
    "py_total": 8,
    "py_week_total": (9, 8),  # row 9, col 8
    "forecast_breakfast": 11,
    "forecast_lunch": 12,
    "forecast_dinner": 13,
    "forecast_total": 14,
    "greeter_breakfast": 18,
    "nonmeal_breakfast": 19,
    "student_breakfast": 20,
    "total_breakfast": 21,
    "greeter_lunch": 24,
    "nonmeal_lunch": 25,
    "student_lunch": 26,
    "total_lunch": 27,
    "greeter_dinner": 30,
    "fulltime_dinner": 31,
    "student_dinner": 32,
    "total_dinner": 33,
    "total_day": 36,
    "week_total": (38, 8),
    "admission": 52,
    "special_groups": 55,
}

# Days map: col 2=Fri, 3=Sat, 4=Sun, 5=Mon, 6=Tue, 7=Wed, 8=Thu
_DAY_COLS = list(range(2, 9))
_DAY_NAMES = ["Fri", "Sat", "Sun", "Mon", "Tue", "Wed", "Thu"]

# Skip these sheet names
_SKIP_SHEETS = {"AVE WEEKLY COUNTS AT GLANCE", "Feuil1"}


def _safe_int(val):
    """Convert cell value to int, handling None, strings, errors."""
    if val is None:
        return 0
    if isinstance(val, str):
        val = val.strip()
        if not val or val.startswith("#") or val == "Select Dropdown":
            return 0
        try:
            return int(float(val.replace(",", "")))
        except ValueError:
            return 0
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return 0


def _get_row_vals(ws, row_num):
    """Get 7 day values from a row (cols B-H)."""
    return [_safe_int(ws.cell(row=row_num, column=c).value) for c in _DAY_COLS]


def _get_dates(ws):
    """Get 7 dates from row 3."""
    dates = []
    for c in _DAY_COLS:
        v = ws.cell(row=3, column=c).value
        if isinstance(v, datetime):
            dates.append(v.strftime("%Y-%m-%d"))
        elif isinstance(v, date):
            dates.append(v.isoformat())
        elif v:
            try:
                dates.append(str(v)[:10])
            except Exception:
                dates.append(None)
        else:
            dates.append(None)
    return dates


def parse_weekly_sheet(ws, sheet_name):
    """
    Parse a single weekly sheet.
    Returns list of daily records (one per day with data).
    """
    dates = _get_dates(ws)
    if not any(dates):
        return []

    # Get all row data
    py_bfast = _get_row_vals(ws, 5)
    py_lunch = _get_row_vals(ws, 6)
    py_dinner = _get_row_vals(ws, 7)

    fc_bfast = _get_row_vals(ws, 11)
    fc_lunch = _get_row_vals(ws, 12)
    fc_dinner = _get_row_vals(ws, 13)

    greeter_bfast = _get_row_vals(ws, 18)
    nonmeal_bfast = _get_row_vals(ws, 19)
    student_bfast = _get_row_vals(ws, 20)
    total_bfast = _get_row_vals(ws, 21)

    greeter_lunch = _get_row_vals(ws, 24)
    nonmeal_lunch = _get_row_vals(ws, 25)
    student_lunch = _get_row_vals(ws, 26)
    total_lunch = _get_row_vals(ws, 27)

    greeter_dinner = _get_row_vals(ws, 30)
    fulltime_dinner = _get_row_vals(ws, 31)
    student_dinner = _get_row_vals(ws, 32)
    total_dinner = _get_row_vals(ws, 33)

    total_day = _get_row_vals(ws, 36)

    admission = _get_row_vals(ws, 52)
    special = _get_row_vals(ws, 55)

    # Weather/notes (rows 42-48)
    weather = []
    notes = []
    for i, r in enumerate(range(42, 49)):
        w = ws.cell(row=r, column=2).value
        n = ws.cell(row=r, column=3).value
        weather.append(str(w) if w and str(w) != "Select Dropdown" else "")
        notes.append(str(n)[:200] if n else "")

    week_total = _safe_int(ws.cell(row=38, column=8).value)

    records = []
    for i in range(7):
        if not dates[i]:
            continue
        records.append({
            "date": dates[i],
            "day_name": _DAY_NAMES[i],
            "sheet_name": sheet_name,
            "location": "Board",
            # Prior year
            "py_breakfast": py_bfast[i],
            "py_lunch": py_lunch[i],
            "py_dinner": py_dinner[i],
            # Forecast
            "fc_breakfast": fc_bfast[i],
            "fc_lunch": fc_lunch[i],
            "fc_dinner": fc_dinner[i],
            # Actual breakdown
            "greeter_breakfast": greeter_bfast[i],
            "nonmeal_breakfast": nonmeal_bfast[i],
            "student_breakfast": student_bfast[i],
            "total_breakfast": total_bfast[i],
            "greeter_lunch": greeter_lunch[i],
            "nonmeal_lunch": nonmeal_lunch[i],
            "student_lunch": student_lunch[i],
            "total_lunch": total_lunch[i],
            "greeter_dinner": greeter_dinner[i],
            "fulltime_dinner": fulltime_dinner[i],
            "student_dinner": student_dinner[i],
            "total_dinner": total_dinner[i],
            "total_day": total_day[i],
            # Groups
            "admission": admission[i],
            "special_groups": special[i],
            # Weather/notes
            "weather": weather[i],
            "notes": notes[i],
        })

    return records


def parse_all_weeks(excel_path):
    """Parse all weekly sheets from the Excel file."""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    all_records = []

    for name in wb.sheetnames:
        if name in _SKIP_SHEETS:
            continue
        # Skip average sheets
        if "average" in name.lower() or "ave " in name.lower():
            continue

        ws = wb[name]
        records = parse_weekly_sheet(ws, name)
        all_records.extend(records)

    return all_records


# ─── Database ───

def init_meal_counts_tables(conn):
    """Create meal counts table."""
    conn.execute("""
        CREATE TABLE IF NOT EXISTS digital_meal_counts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT NOT NULL,
            day_name TEXT,
            sheet_name TEXT,
            location TEXT DEFAULT 'Board',
            py_breakfast INTEGER DEFAULT 0,
            py_lunch INTEGER DEFAULT 0,
            py_dinner INTEGER DEFAULT 0,
            fc_breakfast INTEGER DEFAULT 0,
            fc_lunch INTEGER DEFAULT 0,
            fc_dinner INTEGER DEFAULT 0,
            greeter_breakfast INTEGER DEFAULT 0,
            nonmeal_breakfast INTEGER DEFAULT 0,
            student_breakfast INTEGER DEFAULT 0,
            total_breakfast INTEGER DEFAULT 0,
            greeter_lunch INTEGER DEFAULT 0,
            nonmeal_lunch INTEGER DEFAULT 0,
            student_lunch INTEGER DEFAULT 0,
            total_lunch INTEGER DEFAULT 0,
            greeter_dinner INTEGER DEFAULT 0,
            fulltime_dinner INTEGER DEFAULT 0,
            student_dinner INTEGER DEFAULT 0,
            total_dinner INTEGER DEFAULT 0,
            total_day INTEGER DEFAULT 0,
            admission INTEGER DEFAULT 0,
            special_groups INTEGER DEFAULT 0,
            weather TEXT DEFAULT '',
            notes TEXT DEFAULT '',
            imported_at TEXT DEFAULT (datetime('now','localtime')),
            UNIQUE(date, location)
        )
    """)
    conn.commit()


def import_meal_counts(conn, records):
    """Import parsed records into database."""
    init_meal_counts_tables(conn)
    count = 0
    for r in records:
        # Skip days with no data at all
        if r["total_day"] == 0 and r["total_breakfast"] == 0 and r["total_lunch"] == 0 and r["total_dinner"] == 0:
            continue
        conn.execute("""
            INSERT OR REPLACE INTO digital_meal_counts
            (date, day_name, sheet_name, location,
             py_breakfast, py_lunch, py_dinner,
             fc_breakfast, fc_lunch, fc_dinner,
             greeter_breakfast, nonmeal_breakfast, student_breakfast, total_breakfast,
             greeter_lunch, nonmeal_lunch, student_lunch, total_lunch,
             greeter_dinner, fulltime_dinner, student_dinner, total_dinner,
             total_day, admission, special_groups, weather, notes)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            r["date"], r["day_name"], r["sheet_name"], r["location"],
            r["py_breakfast"], r["py_lunch"], r["py_dinner"],
            r["fc_breakfast"], r["fc_lunch"], r["fc_dinner"],
            r["greeter_breakfast"], r["nonmeal_breakfast"],
            r["student_breakfast"], r["total_breakfast"],
            r["greeter_lunch"], r["nonmeal_lunch"],
            r["student_lunch"], r["total_lunch"],
            r["greeter_dinner"], r["fulltime_dinner"],
            r["student_dinner"], r["total_dinner"],
            r["total_day"], r["admission"], r["special_groups"],
            r["weather"], r["notes"],
        ))
        count += 1
    conn.commit()
    return count


def get_weekly_summary(conn, location="Board"):
    """Get weekly summaries for the dashboard."""
    rows = conn.execute("""
        SELECT
            MIN(date) as week_start,
            MAX(date) as week_end,
            sheet_name,
            SUM(total_breakfast) as breakfast,
            SUM(total_lunch) as lunch,
            SUM(total_dinner) as dinner,
            SUM(total_day) as total,
            SUM(admission) as admission,
            SUM(special_groups) as special
        FROM digital_meal_counts
        WHERE location = ?
        GROUP BY sheet_name
        ORDER BY MIN(date)
    """, (location,)).fetchall()
    return [dict(r) for r in rows]


def get_daily_detail(conn, week_start, week_end, location="Board"):
    """Get daily detail for a specific week."""
    rows = conn.execute("""
        SELECT * FROM digital_meal_counts
        WHERE location = ? AND date >= ? AND date <= ?
        ORDER BY date
    """, (location, week_start, week_end)).fetchall()
    return [dict(r) for r in rows]


if __name__ == "__main__":
    import sqlite3
    conn = sqlite3.connect("budget.db")
    conn.row_factory = sqlite3.Row

    excel_path = "DIGITAL MEAL COUNTS 25.26.xlsx"
    print("Parsing {}...".format(excel_path))
    records = parse_all_weeks(excel_path)
    print("Found {} daily records".format(len(records)))

    count = import_meal_counts(conn, records)
    print("Imported {} records to database".format(count))

    # Show summary
    summary = get_weekly_summary(conn)
    print("\nWeekly Summary:")
    for s in summary:
        print("  {} → {} | B:{} L:{} D:{} | Total: {}".format(
            s["week_start"], s["week_end"],
            s["breakfast"], s["lunch"], s["dinner"], s["total"]))
